# services/mdl_builder.py
import re
import os
import html
import json
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple

import requests

from utils.text_utils import short_text, norm_ref, title_with_acronyms, allcaps, with_The_allcaps, with_the_allcaps
from mdl_helpers import finding_summaries_list

logging.basicConfig(level=logging.INFO)


def format_letter_date(date_iso: Optional[str] = None) -> Tuple[str, str]:
    dt = datetime.fromisoformat(date_iso) if date_iso else datetime.utcnow()
    return dt.strftime("%Y-%m-%d"), dt.strftime("%B %d, %Y")


def summarize_finding_text(raw: str, max_chars: int = 1000) -> str:
    if not raw:
        return ""
    text = re.sub(r"\s+", " ", raw).strip()
    parts = re.split(r"(?<=[.?!])\s+", text)
    picked = []
    for p in parts:
        if len(picked) >= 3:
            break
        if re.search(r"\b(Assistance Listing|Award Period|Federal Program|Identification Number|CFDA)\b", p, re.I):
            continue
        picked.append(p)
    out = " ".join(picked) or text
    return short_text(out, max_chars)


def load_finding_mappings(xlsx_path: Optional[str]):
    """
    Returns:
      - type_map: {'I': 'Procurement and suspension and debarment', ...}
      - summary_labels: ['Lack of evidence of suspension and debarment verification', ...]
    Tolerant to header naming; no-op if workbook missing.
    """
    type_map, summary_labels = {}, []
    if not xlsx_path:
        return type_map, summary_labels
    try:
        import openpyxl
    except Exception:
        return type_map, summary_labels

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception:
        return type_map, summary_labels

    def _find_sheet(*names):
        for ws in wb.worksheets:
            nm = (ws.title or "").strip().lower().replace(" ", "_")
            for want in names:
                if want in nm:
                    return ws
        return None

    # 1) Finding Types sheet
    ws_types = _find_sheet("finding_types", "findingtype", "finding_types_sheet", "types")
    if ws_types and ws_types.max_row >= 2:
        hdrs = [(c.value or "") for c in ws_types[1]]
        hl = [str(h).strip().lower() for h in hdrs]

        def colidx(cands):
            for i, h in enumerate(hl):
                for c in cands:
                    c = c.lower()
                    if h == c or c in h:
                        return i
            return None

        i_code = colidx(["code", "compliance type", "compliance_type", "ctype"])
        i_name = colidx(["name", "description", "label", "type name", "type"])
        if i_code is not None and i_name is not None:
            for row in ws_types.iter_rows(min_row=2, values_only=True):
                code = (row[i_code] or "")
                name = (row[i_name] or "")
                code = str(code).strip().upper()
                name = str(name).strip()
                if code and name:
                    type_map[code] = name

    # 2) Finding_summaries sheet
    ws_summ = _find_sheet("finding_summaries", "finding_summ", "summaries", "summary")
    if ws_summ and ws_summ.max_row >= 2:
        hdrs = [(c.value or "") for c in ws_summ[1]]
        hl = [str(h).strip().lower() for h in hdrs]

        def cidx(cands):
            for i, h in enumerate(hl):
                for c in cands:
                    c = c.lower()
                    if h == c or c in h:
                        return i
            return None

        i_lbl = cidx(["summary", "label", "finding summary", "finding_label"]) or 0
        for row in ws_summ.iter_rows(min_row=2, values_only=True):
            cell = row[i_lbl] if i_lbl < len(row) else None
            txt = str(cell or "").strip()
            if txt:
                summary_labels.append(txt)

    return type_map, summary_labels


def best_summary_label(summary: str, labels: List[str]) -> Optional[str]:
    """
    Offline fuzzy match: pick the label with highest similarity to the summary.
    """
    if not summary or not labels:
        return None
    import difflib
    cand = difflib.get_close_matches(summary, labels, n=1, cutoff=0.0)
    if cand:
        return cand[0]
    return None


def best_summary_label_openai(summary: str, labels: List[str]) -> Optional[str]:
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key or not labels:
        return None

    # Create a clear prompt for matching
    prompt = (
        f"Given the following audit finding text, select the SINGLE best matching category from the list below.\n\n"
        f"Finding text:\n{summary}\n\n"
        f"Categories:\n" + "\n".join(f"- {label}" for label in labels) + "\n\n"
        f"Respond with ONLY the exact category text from the list above that best matches this finding."
    )

    try:
        r = requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            data=json.dumps({
                "model": "gpt-4o-mini",
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0
            }),
            timeout=12,
        )
        out = r.json()
        txt = (out.get("choices", [{}])[0].get("message", {}).get("content") or "").strip()
        if txt in labels:
            return txt
    except Exception:
        pass
    return None


def build_mdl_model_from_fac(
    *,
    auditee_name: str,
    ein: str,
    audit_year: int,
    fac_general: List[Dict[str, Any]],
    fac_findings: List[Dict[str, Any]],
    fac_findings_text: List[Dict[str, Any]],
    fac_caps: List[Dict[str, Any]],
    federal_awards: List[Dict[str, Any]],
    period_end_text: Optional[str] = None,
    address_lines: Optional[List[str]] = None,
    attention_line: Optional[str] = None,
    only_flagged: bool = False,
    max_refs: int = 10,
    auto_cap_determination: bool = True,
    include_no_qc_line: bool = False,
    include_no_cap_line: bool = False,
    treasury_listings: Optional[List[str]] = None,
    aln_reference_xlsx: Optional[str] = None,
    aln_overrides_by_finding: Optional[Dict[str, str]] = None,
    **_
) -> Dict[str, Any]:
    """
    Builds the MDL model.
    """
    # ========== LOAD MAPPINGS ONCE AT THE TOP ==========
    # Default hardcoded mappings as fallback
    type_map = {
        "A": "Activities allowed or unallowed",
        "B": "Allowable costs/cost principles",
        "C": "Cash management",
        "E": "Eligibility",
        "F": "Equipment and real property management",
        "G": "Matching, level of effort, earmarking",
        "H": "Period of performance (or availability) of Federal funds",
        "I": "Procurement and suspension and debarment",
        "J": "Program income",
        "L": "Reporting",
        "M": "Subrecipient monitoring",
        "N": "Special tests and provisions",
        "P": "Other"
    }
    summary_labels = finding_summaries_list  # from mdl_helpers.py

    # Try to load from Excel (only if path provided and file exists)
    if aln_reference_xlsx:
        loaded_type_map, loaded_summary_labels = load_finding_mappings(aln_reference_xlsx)
        if loaded_type_map:  # Only override if we got data
            type_map = loaded_type_map
            logging.info(f"Loaded type_map from Excel with {len(loaded_type_map)} entries")
        if loaded_summary_labels:
            summary_labels = loaded_summary_labels
            logging.info(f"Loaded {len(loaded_summary_labels)} summary labels from Excel")

    logging.info(f"Final type_map: {type_map}")
    logging.info(f"Looking up 'I': {type_map.get('I')}")

    # --------- helpers ----------
    def _derive_assistance_listing(program_name: str, fallback: str = "Unknown") -> str:
        m = re.search(r"\b\d{2}\.\d{3}\b", program_name or "")
        return m.group(0) if m else fallback

    def _title_with_acronyms_inner(s: str) -> str:
        """Title-case but preserve ALL-CAPS tokens (e.g., SLFRF) and common stop words."""
        if not s:
            return ""
        lowers = {"and", "or", "the", "of", "for", "to", "in", "on", "by", "with", "a", "an"}
        out = []
        for tok in str(s).split():
            if tok.isupper() and len(tok) > 1:
                out.append(tok)  # keep acronym
            else:
                w = tok.lower()
                out.append(w if w in lowers else w.capitalize())
        return " ".join(out)

    def _load_aln_mapping(xlsx_path: Optional[str]):
        """
        Returns:
          aln_to_label: {'21.027': 'Coronavirus State and Local Fiscal Recovery Funds (SLFRF)', ...}
          name_to_aln:  {'coronavirus state and local fiscal recovery funds': '21.027', ...}
        """
        aln_to_label, name_to_aln = {}, {}
        if not xlsx_path:
            return aln_to_label, name_to_aln
        try:
            import openpyxl
        except Exception:
            return aln_to_label, name_to_aln
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        except Exception:
            return aln_to_label, name_to_aln

        # Heuristic: find the first sheet that looks like a mapping
        for ws in wb.worksheets:
            hdrs = [(c.value.strip() if isinstance(c.value, str) else (c.value or "")) for c in ws[1]]
            if not any(hdrs):
                continue
            hl = [str(h).strip().lower() for h in hdrs]

            def _find_col(candidates: List[str]) -> Optional[int]:
                for i, h in enumerate(hl):
                    for cand in candidates:
                        cand = cand.lower()
                        if h == cand or cand in h:
                            return i
                return None

            i_aln = _find_col(["aln", "assistance listing", "assistance listing number", "cfda", "cfda number"])
            i_prog = _find_col(["program", "program name", "assistance listing title", "program title"])
            i_acr = _find_col(["acronym", "short", "short name", "abbrev", "abbreviation"])

            if i_prog is None or (i_aln is None and i_acr is None):
                continue  # not a mapping sheet

            for row in ws.iter_rows(min_row=2, values_only=True):
                raw_aln = (row[i_aln] if i_aln is not None else "") or ""
                raw_prog = (row[i_prog] or "")
                raw_acr = (row[i_acr] if i_acr is not None else "") or ""

                aln = str(raw_aln).strip()
                prog_name = str(raw_prog).strip()
                acr = str(raw_acr).strip()
                if not prog_name:
                    continue

                canonical_name = _title_with_acronyms_inner(prog_name)
                if acr:
                    canonical_name = f"{canonical_name} ({acr})"

                if aln:
                    aln_to_label[aln] = canonical_name
                name_to_aln[prog_name.lower()] = aln  # may be "" if not provided
            break  # first matching sheet is enough

        return aln_to_label, name_to_aln

    def _apply_canonicalization_after_grouping(
        group: Dict[str, Any],
        aln_to_label: Dict[str, str],
        name_to_aln: Dict[str, str],
    ):
        """
        Normalize 'assistance_listing' and 'program_name' in-place for a group.
        """
        cur_aln = (group.get("assistance_listing") or "").strip()
        cur_name = (group.get("program_name") or "").strip()

        # 1) If we already have a valid ALN, use canonical label
        if cur_aln and cur_aln in aln_to_label:
            group["assistance_listing"] = cur_aln
            group["program_name"] = aln_to_label[cur_aln]
            return

        # 2) If ALN missing/Unknown, try via name -> ALN
        guess_aln = name_to_aln.get(cur_name.lower())
        if (not cur_aln or cur_aln == "Unknown") and guess_aln:
            group["assistance_listing"] = guess_aln
            group["program_name"] = aln_to_label.get(guess_aln, _title_with_acronyms_inner(cur_name or "Unknown Program"))
            return

        # 3) Treasury heuristics (common programs) - last resort
        nm = cur_name.lower()
        heuristics = [
            ("slfrf", ("21.027", "Coronavirus State and Local Fiscal Recovery Funds (SLFRF)")),
            ("fiscal recovery", ("21.027", "Coronavirus State and Local Fiscal Recovery Funds (SLFRF)")),
            ("emergency rental assistance", ("21.023", "Emergency Rental Assistance Program (ERA)")),
            ("homeowner assistance", ("21.026", "Homeowner Assistance Fund (HAF)")),
            ("capital projects fund", ("21.029", "Capital Projects Fund (CPF)")),
            ("state small business credit", ("21.031", "State Small Business Credit Initiative (SSBCI)")),
            ("local assistance and tribal consistency", ("21.032", "Local Assistance and Tribal Consistency Fund (LATCF)")),
        ]
        for key, (aln_guess, label) in heuristics:
            if key in nm:
                group["assistance_listing"] = aln_guess
                group["program_name"] = label
                break

        # 4) Final tidy for program name casing if still raw/all-caps
        final_name = (group.get("program_name") or "").strip()
        if final_name.isupper() or not final_name:
            group["program_name"] = _title_with_acronyms_inner(final_name or cur_name or "Unknown Program")
        if not group.get("assistance_listing"):
            group["assistance_listing"] = "Unknown"

    # --------- load mapping once ----------
    aln_to_label, name_to_aln = _load_aln_mapping(aln_reference_xlsx)
    logging.info(f"Loaded type_map: {type_map}")
    logging.info(f"Loaded {len(summary_labels)} summary labels")

    # --------- award lookups from FAC ----------
    award2meta: Dict[str, Dict[str, str]] = {}
    for a in (federal_awards or []):
        ref = a.get("award_reference")
        pname = (a.get("federal_program_name") or "").strip()
        explicit_aln = (a.get("assistance_listing") or a.get("assistance_listing_number") or "").strip()
        derived_aln = _derive_assistance_listing(pname, fallback="")
        aln = explicit_aln or derived_aln or "Unknown"

        # Prefer Excel canonical if we have the ALN
        if aln != "Unknown" and aln in aln_to_label:
            canonical_name = aln_to_label[aln]
        else:
            # Try map by name -> ALN
            mapped_aln = name_to_aln.get(pname.lower())
            if mapped_aln:
                aln = mapped_aln
                canonical_name = aln_to_label.get(mapped_aln, _title_with_acronyms_inner(pname or "Unknown Program"))
            else:
                canonical_name = _title_with_acronyms_inner(pname or "Unknown Program")

        if ref:
            award2meta[ref] = {
                "program_name": canonical_name or "Unknown Program",
                "assistance_listing": aln or "Unknown",
            }

    logging.info(f"Built award2meta with {len(award2meta)} entries:")
    for k, v in award2meta.items():
        logging.info(f" AWARDS {k}: {v}")

    # --------- text / CAP lookups ----------
    text_by_ref = {
        norm_ref(t.get("finding_ref_number")): (t.get("finding_text") or "").strip()
        for t in (fac_findings_text or [])
    }
    cap_by_ref = {
        norm_ref(c.get("finding_ref_number")): (c.get("planned_action") or "").strip()
        for c in (fac_caps or [])
    }

    def _is_flagged(f: dict) -> bool:
        return any([
            f.get("is_material_weakness") is True,
            f.get("is_significant_deficiency") is True,
            f.get("is_questioned_costs") is True,
            f.get("is_modified_opinion") is True,
            f.get("is_other_findings") is True,
            f.get("is_other_matters") is True,
            f.get("is_repeat_finding") is True,
        ])

    # --------- choose finding refs ----------
    base_refs: List[str] = []
    for f in (fac_findings or []):
        if only_flagged and not _is_flagged(f):
            continue
        r = f.get("reference_number")
        if r:
            base_refs.append(r)

    if not base_refs and fac_findings_text:
        base_refs = [t.get("finding_ref_number") for t in fac_findings_text if t.get("finding_ref_number")]

    seen = set()
    norm_refs: List[Tuple[str, str]] = []
    for r in base_refs:
        if not r:
            continue
        k = norm_ref(r)
        if k not in seen:
            seen.add(k)
            norm_refs.append((r, k))
    norm_refs = norm_refs[: max_refs or 10]
    chosen_keys = {kn for _, kn in norm_refs}

    # --------- group findings under award_reference ----------
    programs_map: Dict[str, Dict[str, Any]] = {}
    for f in (fac_findings or []):
        r = f.get("reference_number")
        if not r:
            continue
        k = norm_ref(r)
        if k not in chosen_keys:
            continue

        award_ref = f.get("award_reference") or "UNKNOWN"
        logging.info(f"Finding {r} -> award_ref: {award_ref}")

        # Try to get metadata from award lookup
        meta = award2meta.get(award_ref, {})

        # If not found in award2meta, try aln_overrides_by_finding
        if not meta.get("assistance_listing") or meta.get("assistance_listing") == "Unknown":
            if aln_overrides_by_finding and r in aln_overrides_by_finding:
                override_aln = aln_overrides_by_finding[r]
                logging.info(f"   Using finding-level ALN override: {override_aln}")
                meta["assistance_listing"] = override_aln
                # Update program name if we have ALN mapping
                if override_aln in aln_to_label:
                    meta["program_name"] = aln_to_label[override_aln]

        logging.info(f"   Final meta: {meta}")

        group = programs_map.setdefault(award_ref, {
            "assistance_listing": meta.get("assistance_listing", "Unknown"),
            "program_name": meta.get("program_name", "Unknown Program"),
            "findings": []
        })

        # If ALN is Unknown, try to fill from finding-level override (XLSX)
        if group.get("assistance_listing") in (None, "", "Unknown"):
            orig_ref = f.get("reference_number") or ""
            cand_aln = None
            if aln_overrides_by_finding:
                cand_aln = (aln_overrides_by_finding.get(orig_ref)
                            or aln_overrides_by_finding.get(norm_ref(orig_ref)))
            if cand_aln:
                group["assistance_listing"] = cand_aln
                if cand_aln in aln_to_label:
                    group["program_name"] = aln_to_label[cand_aln]

        summary = summarize_finding_text(text_by_ref.get(k, ""))
        cap_text = cap_by_ref.get(k)

        qcost_det = "Questioned Cost:\nNone\nDisallowed Cost:\nNone" if include_no_qc_line else "None"
        cap_det = (
            "Accepted" if (auto_cap_determination and cap_text)
            else ("No CAP required" if include_no_cap_line else "Not Applicable")
        )

        ctype_code = (f.get("type_requirement") or "").strip().upper()[:1]
        ctype_label = type_map.get(ctype_code) or ctype_code or ""

        # Get the complete finding summary text
        complete_summary = text_by_ref.get(k, "")

        # Match against standardized summaries - try OpenAI first with COMPLETE text
        matched_label = None
        if complete_summary:
            matched_label = (best_summary_label_openai(complete_summary, summary_labels)
                           or best_summary_label(complete_summary, summary_labels))

        # Fallback to shortened summary if no match
        if not matched_label:
            matched_label = summary

        logging.info(f"Finding {f.get('reference_number')}: {ctype_label} - {matched_label}")
        logging.info(f"Matched label: {matched_label}")
        logging.info(f"Compliance type: {ctype_label}")
        logging.info(f"Summary: {summary}")
        logging.info(f" {ctype_label}, {summary}, {cap_text}, {qcost_det}, {cap_det}")

        group["findings"].append({
            "finding_id": f.get("reference_number") or "",
            "compliance_type": ctype_label,  # Full label: "Procurement and suspension and debarment"
            "summary": matched_label,  # Matched standardized summary
            "compliance_and_summary": f"{ctype_label} - {matched_label}".strip(" -"),  # Combined for display
            "audit_determination": "Sustained",
            "questioned_cost_determination": qcost_det,
            "disallowed_cost_determination": "None",
            "cap_determination": cap_det,
            "cap_text": cap_text,
            "is_repeat_finding": f.get("is_repeat_finding") is True,
        })

    # If nothing grouped but we have refs, emit a catch-all
    if not programs_map and norm_refs:
        catchall = {"assistance_listing": "Unknown", "program_name": "Unknown Program", "findings": []}

        for orig, key in norm_refs:
            finding_data = None
            for f in (fac_findings or []):
                if norm_ref(f.get("reference_number")) == key:
                    finding_data = f
                    break

            if finding_data:
                ctype_code = (finding_data.get("type_requirement") or "").strip().upper()[:1]
            else:
                ctype_code = ""

            ctype_label = type_map.get(ctype_code) or ctype_code or ""

            complete_summary = text_by_ref.get(key, "")
            matched_label = None
            if complete_summary:
                matched_label = (best_summary_label_openai(complete_summary, summary_labels)
                            or best_summary_label(complete_summary, summary_labels))
            if not matched_label:
                matched_label = summarize_finding_text(complete_summary)

            cap_text = cap_by_ref.get(key)
            qcost_det = "No questioned costs identified" if include_no_qc_line else "None"
            cap_det = (
                "Accepted" if (auto_cap_determination and cap_text)
                else ("No CAP required" if include_no_cap_line else "Not Applicable")
            )

            catchall["findings"].append({
                "finding_id": orig,
                "compliance_type": ctype_label,
                "summary": matched_label,
                "compliance_and_summary": f"{ctype_label} - {matched_label}".strip(" -"),
                "audit_determination": "Sustained",
                "questioned_cost_determination": qcost_det,
                "disallowed_cost_determination": "None",
                "cap_determination": cap_det,
                "cap_text": cap_text,
                "is_repeat_finding": finding_data.get("is_repeat_finding") is True if finding_data else False,
            })
        programs_map["UNKNOWN"] = catchall

    # ----- Canonicalize program fields using the Excel mapping (now that groups exist)
    for grp in programs_map.values():
        _apply_canonicalization_after_grouping(grp, aln_to_label, name_to_aln)

    # ----- Apply Treasury ALN filter AFTER canonicalization
    if treasury_listings:
        allowed = {(aln or "").strip() for aln in treasury_listings if aln}
        logging.info(f" Treasury listings filter: {allowed}")
        logging.info(f"  Programs before filter: {list(programs_map.keys())}")
        programs_map = {k: v for k, v in programs_map.items() if v.get("assistance_listing") in allowed}
        logging.info(f" Programs after filter: {list(programs_map.keys())}")

    # ----- Build final model
    model = {
        "letter_date_iso": datetime.utcnow().strftime("%Y-%m-%d"),
        "auditee_name": auditee_name,
        "ein": f"{ein[:2]}-{ein[2:]}" if ein and ein.isdigit() and len(ein) == 9 else ein,
        "address_lines": address_lines or [],
        "attention_line": attention_line or "",
        "period_end_text": period_end_text or f"June 30, {audit_year}",
        "audit_year": audit_year,
        "programs": list(programs_map.values()),
        "not_sustained_notes": [],
    }
    return model


def render_mdl_html(model: Dict[str, Any]) -> str:
    letter_date_iso = model.get("letter_date_iso")
    _, letter_date_long = format_letter_date(letter_date_iso)

    auditee_name = model.get("auditee_name", "Recipient")
    ein = model.get("ein", "")
    address_lines = model.get("address_lines", [])
    attention_line = model.get("attention_line")
    period_end_text = model.get("period_end_text", str(model.get("audit_year", "")))
    include_no_qc_line = model.get("include_no_qc_line", True)
    treasury_contact_email = model.get("treasury_contact_email", "ORP_SingleAudits@treasury.gov ")
    address_block = "<br>".join(html.escape(x) for x in address_lines) if address_lines else ""
    attention_block = f"<p><strong>{html.escape(attention_line)}</strong></p>" if attention_line else ""

    from mdl_helpers import _combine_comp_summary

    def _render_program_table(p: Dict[str, Any]) -> str:
        rows_html = []
        for f in p.get("findings", []):
            repeat = "Yes" if f.get("is_repeat_finding") else "No"
            rows_html.append(f"""
              <tr>
                <td>{html.escape(f.get('finding_id',''))}</td>
                <td>{html.escape(_combine_comp_summary(f))}</td>
                <td>{html.escape(f.get('audit_determination',''))}</td>
                <td>{html.escape(f.get('questioned_cost_determination',''))}</td>
                <td>{html.escape(f.get('cap_determination',''))}</td>
                <td>{html.escape(repeat)}</td>
              </tr>
            """)
        if not rows_html:
            rows_html.append("""
              <tr><td colspan="6"><em>No MDL-relevant findings identified for this program.</em></td></tr>
            """)
        table = f"""
          <table border="1" cellspacing="0" cellpadding="6" style="border-collapse:collapse; width:100%; font-size:10.5pt;">
            <tr>
              <th>Audit<br>Finding #</th>
              <th>Compliance Type -<br>Audit Finding Summary</th>
              <th>Audit Finding<br>Determination</th>
              <th>Questioned Cost<br>Determination</th>
              <th>CAP<br>Determination</th>
              <th>Repeat<br>Finding</th>
            </tr>
            {''.join(rows_html)}
          </table>
        """
        cap_blocks = []
        for f in p.get("findings", []):
            cap_text = f.get("cap_text")
            if cap_text:
                cap_blocks.append(f"""
                  <h4>Corrective Action Plan – {html.escape(f.get('finding_id',''))}</h4>
                  <p>{html.escape(cap_text)}</p>
                """)
        return table + ("\n".join(cap_blocks) if cap_blocks else "")

    programs = model.get("programs", [])
    programs_html = "\n".join(_render_program_table(p) for p in programs) if programs else "<p><em>No MDL-relevant findings identified per FAC records.</em></p>"

    not_sustained_notes = model.get("not_sustained_notes", [])
    not_sustained_html = ""
    if not_sustained_notes:
        notes_paras = "\n".join(f"<p>{html.escape(n)}</p>" for n in not_sustained_notes if n)
        not_sustained_html = f"<h3>FINDINGS NOT SUSTAINED</h3>\n{notes_paras}"

    chunks = []
    chunks.append(f'<p style="text-align:right; margin:0 0 12pt 0;">{html.escape(letter_date_long)}</p>')
    chunks.append("""
      <p style="margin:0 0 12pt 0;">
        <strong>DEPARTMENT OF THE TREASURY</strong><br>
        WASHINGTON, D.C.
      </p>
    """)
    chunks.append(f"""
      <p style="margin:0 0 12pt 0;">
        <strong>{html.escape(allcaps(auditee_name))}</strong><br>
        EIN: {html.escape(ein)}<br>
        {address_block}
      </p>
    """)
    if attention_block:
        chunks.append(attention_block)
    chunks.append(f"""
      <p style="margin:12pt 0 12pt 0;">
        <strong>Subject:</strong> U.S. Department of the Treasury's Management Decision Letter (MDL) for Single Audit Report for the period ending on {html.escape(period_end_text)}
      </p>
    """)
    chunks.append("""
      <p>
        In accordance with 2 C.F.R. § 200.521(b), the U.S. Department of the Treasury (Treasury)
        is required to issue a management decision for single audit findings pertaining to awards under
        Treasury's programs. Treasury's review as part of its responsibilities under 2 C.F.R. § 200.513(c)
        includes an assessment of Treasury's award recipients' single audit findings, corrective action plans (CAPs),
        and questioned costs, if any.
      </p>
    """)
    chunks.append(f"""
    <p>
        Treasury has reviewed the single audit report for {html.escape(with_The_allcaps(auditee_name))},
        prepared by {html.escape(with_the_allcaps(model.get("auditor_name","")))} for the fiscal year ending on
        {html.escape(period_end_text)}.
        Treasury has made the following determinations regarding the audit finding(s) and CAP(s) listed below.
    </p>
    """)
    if include_no_qc_line:
        chunks.append("<p>No questioned costs are included in this single audit report.</p>")

    chunks.append(programs_html)
    if not_sustained_html:
        chunks.append(not_sustained_html)

    chunks.append("""
      <p>
        Please note, the corrective action(s) are subject to review during the recipient's next annual single audit
        or program-specific audit, as applicable, to determine adequacy. If the same audit finding(s) appear in a future single
        audit report for this recipient, its current or future award funding under Treasury's programs may be adversely impacted.
      </p>
      <p>
        For questions regarding the audit finding(s), please email us at
        <a href="mailto:{html.escape(treasury_contact_email)}">{html.escape(treasury_contact_email)}</a>.
        Thank you.
    </p>
      <p style="margin-top:18pt;">Sincerely,<br><br>
      Audit and Compliance Resolution Team<br>
      Office of Capital Access<br>
      U.S. Department of the Treasury</p>
    """)

    return f'<div style="font-family: Calibri, Arial, sans-serif; font-size:11pt; line-height:1.4;">{"".join(chunks)}</div>'
