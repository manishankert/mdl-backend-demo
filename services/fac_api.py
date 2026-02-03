# services/fac_api.py
import os
import re
import logging
from io import BytesIO
from typing import Dict, List, Any

import requests
from fastapi import HTTPException

from config import FAC_BASE, FAC_KEY

logging.basicConfig(level=logging.INFO)


def fac_headers():
    key = os.getenv("FAC_API_KEY") or FAC_KEY
    if not key:
        raise HTTPException(500, "FAC_API_KEY not configured on the docx service")
    return {"X-Api-Key": key}


def fac_get(path: str, params: Dict[str, Any]) -> Any:
    try:
        r = requests.get(f"{FAC_BASE.rstrip('/')}/{path.lstrip('/')}",
                         headers=fac_headers(), params=params, timeout=20)
        r.raise_for_status()
        return r.json()
    except requests.HTTPError as e:
        raise HTTPException(r.status_code if 'r' in locals() else 500,
                            f"FAC GET {path} failed: {getattr(r,'text','')}") from e
    except Exception as e:
        raise HTTPException(500, f"FAC GET {path} failed: {e}") from e


def or_param(field: str, values: List[str]) -> str:
    inner = ",".join([f"{field}.eq.{v}" for v in values])
    return f"({inner})"


def read_headers(ws):
    return [(c.value or "").strip() if isinstance(c.value, str) else (c.value or "") for c in ws[1]]


def find_col(headers, candidates):
    hl = [str(h).strip().lower() for h in headers]
    for i, h in enumerate(hl):
        for cand in candidates:
            cl = cand.strip().lower()
            if h == cl or cl in h:
                return i
    return None


def aln_overrides_from_summary(report_id: str):
    """
    Returns (aln_by_award, aln_by_finding) by parsing the public FAC summary XLSX.
    Updated to handle the actual FAC Excel structure correctly.
    """
    url = f"https://app.fac.gov/dissemination/summary-report/xlsx/{report_id}"
    logging.info(f"Downloading FAC summary from: {url}")

    r = requests.get(url, timeout=20)
    r.raise_for_status()

    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(r.content), data_only=True)

    aln_by_award = {}
    aln_by_finding = {}

    logging.info(f"Excel sheets available: {wb.sheetnames}")

    # ============================================================
    # PART 1: Process FEDERALAWARD sheet
    # ============================================================
    if 'federalaward' in wb.sheetnames:
        ws_fed = wb['federalaward']
        logging.info(f"\nProcessing 'federalaward' sheet (range: {ws_fed.dimensions})")

        # Read headers from row 1
        headers = [str(cell.value or "").strip().lower() for cell in ws_fed[1]]
        logging.info(f"   Headers: {headers}")

        # Find column indices
        try:
            i_award_ref = headers.index('award_reference')
            i_aln = headers.index('aln')
            i_program = headers.index('federal_program_name')

            logging.info(f"   Column indices - award_ref:{i_award_ref}, aln:{i_aln}, program:{i_program}")

            # Process data rows (starting from row 2)
            award_count = 0
            for row in ws_fed.iter_rows(min_row=2, values_only=True):
                if not row or all(c is None for c in row):
                    continue

                award_ref = str(row[i_award_ref] or "").strip()
                aln = str(row[i_aln] or "").strip()
                program_name = str(row[i_program] or "").strip()

                # Validate ALN format (should be like 21.027)
                if award_ref and aln and re.match(r'^\d{2}\.\d{3}', aln):
                    aln_by_award[award_ref] = aln
                    award_count += 1
                    logging.info(f"   Award: {award_ref} -> {aln} ({program_name[:50]})")

            logging.info(f"   Processed {award_count} federal awards")

        except ValueError as e:
            logging.warning(f"   Could not find required columns in federalaward sheet: {e}")

    # ============================================================
    # PART 2: Process FINDING sheet
    # ============================================================
    if 'finding' in wb.sheetnames:
        ws_find = wb['finding']
        logging.info(f"\nProcessing 'finding' sheet (range: {ws_find.dimensions})")

        # Read headers from row 1
        headers = [str(cell.value or "").strip().lower() for cell in ws_find[1]]
        logging.info(f"   Headers: {headers}")

        # Find column indices
        try:
            i_ref_num = headers.index('reference_number')
            i_aln = headers.index('aln')
            i_award_ref = headers.index('award_reference')

            logging.info(f"   Column indices - ref_num:{i_ref_num}, aln:{i_aln}, award_ref:{i_award_ref}")

            # Process data rows
            finding_count = 0
            for row in ws_find.iter_rows(min_row=2, values_only=True):
                if not row or all(c is None for c in row):
                    continue

                ref_num = str(row[i_ref_num] or "").strip()
                aln = str(row[i_aln] or "").strip()
                award_ref = str(row[i_award_ref] or "").strip()

                # Validate ALN format
                if ref_num and aln and re.match(r'^\d{2}\.\d{3}', aln):
                    aln_by_finding[ref_num] = aln
                    finding_count += 1
                    logging.info(f"   Finding: {ref_num} -> {aln} (Award: {award_ref})")

            logging.info(f"   Processed {finding_count} findings")

        except ValueError as e:
            logging.warning(f"   Could not find required columns in finding sheet: {e}")

    logging.info(f"\nFINAL RESULTS:")
    logging.info(f"   Award mappings: {len(aln_by_award)}")
    logging.info(f"   Finding mappings: {len(aln_by_finding)}")

    if aln_by_award:
        logging.info(f"\n   Sample award mappings:")
        for k, v in list(aln_by_award.items())[:3]:
            logging.info(f"     {k} -> {v}")

    if aln_by_finding:
        logging.info(f"\n   Sample finding mappings:")
        for k, v in list(aln_by_finding.items())[:3]:
            logging.info(f"     {k} -> {v}")

    return aln_by_award, aln_by_finding


def from_fac_general(gen_rows):
    """
    Pulls reasonable defaults from FAC /general for headers.
    Expects select to include:
      auditee_address_line_1,auditee_city,auditee_state,auditee_zip,
      auditor_firm_name,fy_end_date
    """
    if not gen_rows:
        return {}
    g = gen_rows[0]
    # fy_end_date can be 'YYYY-MM-DD' -> convert to 'Month DD, YYYY' if present
    per_end = None
    try:
        from datetime import datetime
        if g.get("fy_end_date"):
            dt = datetime.fromisoformat(g["fy_end_date"])
            per_end = dt.strftime("%B %-d, %Y") if hasattr(dt, "strftime") else None
    except Exception:
        pass

    return {
        "street_address": g.get("auditee_address_line_1") or "",
        "city": g.get("auditee_city") or "",
        "state": g.get("auditee_state") or "",
        "zip_code": g.get("auditee_zip") or "",
        "auditor_name": g.get("auditor_firm_name") or "",
        "period_end_text": per_end,
        "poc_name": g.get("auditee_contact_name") or "",
        "poc_title": g.get("auditee_contact_title") or "",
    }
