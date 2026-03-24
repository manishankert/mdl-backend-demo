"""
bulk_mdl.py
-----------
Reads an Excel sheet with EIN and audit_year columns and calls the
/build-mdl-docx-auto endpoint to produce MDLs in batches of 10.
Downloads each generated DOCX automatically to a local output folder.

Excel format (any sheet name, headers in row 1):
  | ein        | audit_year |
  | 586000804  | 2022       |
  | 046001287  | 2023       |
  ...

Usage:
  python bulk_mdl.py --file input.xlsx --url http://localhost:8000
  python bulk_mdl.py --file input.xlsx --url http://localhost:8000 --output ./my_mdls
"""
import zipfile
import argparse
import time
import logging
from pathlib import Path

import openpyxl
import requests

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

BATCH_SIZE = 10
BATCH_PAUSE_SECONDS = 5
REQUEST_TIMEOUT = 60
DEFAULT_OUTPUT_DIR = "mdl_output"


def load_rows(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    headers = [str(c.value or "").strip().lower() for c in ws[1]]
    logging.info(f"Headers found: {headers}")

    def find_col(candidates):
        for i, h in enumerate(headers):
            for c in candidates:
                if h == c or c in h:
                    return i
        return None

    i_ein = find_col(["ein", "auditee_ein", "employer_identification"])
    i_year = find_col(["audit_year", "year", "fiscal_year"])
    i_name = find_col(["auditee_name", "recipient_name", "name"])

    if i_ein is None or i_year is None:
        raise ValueError(f"Could not find required columns. Headers detected: {headers}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ein = str(row[i_ein] or "").strip().replace("-", "").zfill(9)
        year = row[i_year]
        if not ein or not year:
            continue
        try:
            year = int(year)
        except (ValueError, TypeError):
            logging.warning(f"Skipping row — invalid year: {year}")
            continue
        name = str(row[i_name] or "").strip() if i_name is not None else ""
        rows.append({"ein": ein, "audit_year": year, "auditee_name": name})

    logging.info(f"Loaded {len(rows)} rows from {xlsx_path}")
    return rows


def call_endpoint(base_url: str, ein: str, audit_year: int, auditee_name: str = "", include_sfsac=False) -> dict:
    url = f"{base_url.rstrip('/')}/build-mdl-docx-auto"
    payload = {
        "ein": ein,
        "audit_year": audit_year,
        "auditee_name": auditee_name or None,
        "include_awards": True,
        "only_flagged": False,
        "max_refs": 15,
        "treasury_listings": ["21.032", "21.031", "21.029", "21.027", "21.026", "21.023"],
        "download_sfsac": include_sfsac,
    }
    try:
        r = requests.post(url, json=payload, timeout=REQUEST_TIMEOUT)
        r.raise_for_status()
        return r.json()
    except requests.HTTPError as e:
        return {"ok": False, "message": f"HTTP {r.status_code}: {r.text[:200]}"}
    except Exception as e:
        return {"ok": False, "message": str(e)}


def download_file(url: str, dest_path: Path) -> bool:
    """Download a file from a URL and save it locally. Returns True on success."""
    try:
        r = requests.get(url, timeout=REQUEST_TIMEOUT)
        r.raise_for_status()
        dest_path.write_bytes(r.content)
        logging.info(f"    Downloaded -> {dest_path}")
        return True
    except Exception as e:
        logging.warning(f"    Download failed: {e}")
        return False


'''def run_bulk(xlsx_path: str, base_url: str, output_dir: str, include_sfsac: bool = False):
    rows = load_rows(xlsx_path)
    total = len(rows)
    results = []

    # Create output folder if it doesn't exist
    out_folder = Path(output_dir)
    out_folder.mkdir(parents=True, exist_ok=True)
    logging.info(f"Output folder: {out_folder.resolve()}")

    for batch_start in range(0, total, BATCH_SIZE):
        batch = rows[batch_start: batch_start + BATCH_SIZE]
        batch_num = batch_start // BATCH_SIZE + 1
        logging.info(f"--- Batch {batch_num}: rows {batch_start + 1}–{batch_start + len(batch)} of {total} ---")

        for i, row in enumerate(batch):
            ein = row["ein"]
            year = row["audit_year"]
            logging.info(f"  [{batch_start + i + 1}/{total}] EIN={ein} year={year}")
            result = call_endpoint(base_url, ein, year, auditee_name=row.get("auditee_name", ""), include_sfsac=include_sfsac)

            status = "FAILED"
            url_out = result.get("url", "")
            message = result.get("message", "")
            local_path = ""

            if result.get("ok") and url_out:
                # Build filename matching MDL naming convention:
                # MDL-Recipient_Name-EIN-YEAR.docx
                raw_name = result.get("auditee_name") or result.get("recipient_name") or ""
                if not raw_name:
                    # Try to pull it from blob_path as fallback
                    blob = result.get("blob_path", "")
                    raw_name = blob.split("/")[-1].replace(".docx", "").split("-")[1] if blob else ""
                safe_name = raw_name.strip().replace(" ", "_").replace(",", "").replace(".", "")
                filename = f"MDL-{safe_name}-{ein}-{year}.docx" if safe_name else f"MDL-{ein}-{year}.docx"
                dest = out_folder / filename
                downloaded = download_file(url_out, dest)
                if downloaded:
                    status = "OK"
                    local_path = str(dest.resolve())

                    # Download SF-SAC if returned
                    sfsac_url = result.get("sfsac_url")
                    if sfsac_url:
                        sfsac_filename = f"SF-SAC-{safe_name}-{ein}-{year}.pdf" if safe_name else f"SF-SAC-{ein}-{year}.pdf"
                        sfsac_dest = out_folder / sfsac_filename
                        download_file(sfsac_url, sfsac_dest)
                else:
                    status = "DOWNLOAD_FAILED"
            else:
                logging.warning(f"    FAILED — {message}")

            results.append({
                "ein": ein,
                "audit_year": year,
                "auditee_name": result.get("auditee_name") or result.get("recipient_name") or "",
                "status": status,
                "url": url_out,
                "local_path": local_path,
                "message": message,
            })

        # Pause between batches (skip after last batch)
        if batch_start + BATCH_SIZE < total:
            logging.info(f"  Batch complete. Pausing {BATCH_PAUSE_SECONDS}s before next batch...")
            time.sleep(BATCH_PAUSE_SECONDS)

    # Summary
    ok = sum(1 for r in results if r["status"] == "OK")
    failed = total - ok
    logging.info(f"\n=== DONE: {ok}/{total} succeeded, {failed} failed ===")
    logging.info(f"Files saved to: {out_folder.resolve()}")

    # Write results to Excel
    out_results = Path(xlsx_path).stem + "_results.xlsx"
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Results"
    ws_out.append(["ein", "audit_year", "auditee_name", "status", "local_path", "url", "message"])
    for r in results:
        ws_out.append([r["ein"], r["audit_year"], r["auditee_name"], r["status"], r["local_path"], r["url"], r["message"]])
    wb_out.save(out_results)
    logging.info(f"Results log saved to: {out_results}")

    # Zip everything in the output folder
    zip_name = Path(xlsx_path).stem + ".zip"
    outputs_dir = Path("outputs")
    outputs_dir.mkdir(parents=True, exist_ok=True)
    zip_path = outputs_dir / zip_name
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in out_folder.iterdir():
            if f.is_file():
                zf.write(f, f.name)
    logging.info(f"Zipped all files to: {zip_path.resolve()}")
'''

def run_bulk(xlsx_path: str, base_url: str, output_dir: str, include_sfsac: bool = False):
    rows = load_rows(xlsx_path)
    total = len(rows)

    logging.info(f"Sending {total} items to /build-mdl-docx-bulk...")

    url = f"{base_url.rstrip('/')}/build-mdl-docx-bulk"
    payload = {
        "items": [{"ein": r["ein"], "audit_year": r["audit_year"], "auditee_name": r.get("auditee_name", "")} for r in rows],
        "include_sfsac": include_sfsac,
        "include_awards": True,
        "max_refs": 15,
        "treasury_listings": ["21.032", "21.031", "21.029", "21.027", "21.026", "21.023"],
    }

    try:
        r = requests.post(url, json=payload, timeout=1200)
        r.raise_for_status()
        result = r.json()
    except Exception as e:
        logging.error(f"Bulk request failed: {e}")
        return

    zip_url = result.get("zip_url")
    logging.info(f"Succeeded: {result.get('succeeded')}/{result.get('total')}")

    if zip_url:
        logging.info(f"Zip URL: {zip_url}")
        '''# Download zip to outputs/
        outputs_dir = Path("outputs")
        outputs_dir.mkdir(parents=True, exist_ok=True)
        zip_name = Path(xlsx_path).stem + ".zip"
        zip_path = outputs_dir / zip_name
        try:
            zr = requests.get(zip_url, timeout=120)
            zr.raise_for_status()
            zip_path.write_bytes(zr.content)
            logging.info(f"Zip downloaded to: {zip_path.resolve()}")
        except Exception as e:
            logging.warning(f"Zip download failed: {e}")'''
    else:
        logging.warning("No zip URL returned.")

    # Write results log
    out_results = Path(xlsx_path).stem + "_results.xlsx"
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Results"
    ws_out.append(["ein", "audit_year", "auditee_name", "status", "message"])
    for r in result.get("results", []):
        ws_out.append([r.get("ein"), r.get("audit_year"), r.get("auditee_name", ""), r.get("status"), r.get("message", "")])
    wb_out.save(out_results)
    logging.info(f"Results log saved to: {out_results}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Bulk MDL generator")
    parser.add_argument("--sfsac", action="store_true", help="Also download SF-SAC PDFs")
    parser.add_argument("--file", required=True, help="Path to input Excel file")
    parser.add_argument("--url", required=True, help="Base URL of the MDL service (e.g. http://localhost:8000)")
    parser.add_argument("--output", default=DEFAULT_OUTPUT_DIR, help="Local folder to save downloaded MDLs (default: mdl_output)")
    args = parser.parse_args()

    run_bulk(args.file, args.url, args.output, include_sfsac=args.sfsac)